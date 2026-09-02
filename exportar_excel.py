import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment
import os
from tkinter import messagebox

def configurar_encabezados(hoja, columnas):
    hoja.append(columnas)
    for col in range(1, len(columnas) + 1):
        celda = hoja.cell(row=1, column=col)
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal="center")
        hoja.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

def exportar_inventario(ruta_db, ruta_guardado):
    try:
        conexion = sqlite3.connect(ruta_db)
        cursor = conexion.cursor()
        cursor.execute("SELECT id_producto, nombre, unidad_medida, stock_actual, costo_actual, precio_sugerido FROM productos")
        datos = cursor.fetchall()
        conexion.close()

        wb = openpyxl.Workbook()
        hoja = wb.active
        hoja.title = "Inventario Salsamentaria"

        columnas = ["ID", "Producto", "Unidad", "Stock Actual", "Costo Actual", "Precio Sugerido"]
        configurar_encabezados(hoja, columnas)

        for fila in datos:
            hoja.append(fila)

        wb.save(ruta_guardado)
        os.startfile(ruta_guardado) # Abre el archivo automáticamente al guardarlo
    except Exception as e:
        messagebox.showerror("Error de Exportación", str(e))

def exportar_ventas_dia(ruta_db, ruta_guardado):
    try:
        conexion = sqlite3.connect(ruta_db)
        cursor = conexion.cursor()
        cursor.execute("""
            SELECT numero_factura, datetime(fecha_venta, 'localtime'), total_venta, valor_recibido, vueltas
            FROM ventas 
            WHERE date(fecha_venta, 'localtime') = date('now', 'localtime')
        """)
        datos = cursor.fetchall()
        conexion.close()

        wb = openpyxl.Workbook()
        hoja = wb.active
        hoja.title = "Ventas del Día"

        columnas = ["No. Factura", "Fecha y Hora", "Total Venta", "Valor Recibido", "Vueltas"]
        configurar_encabezados(hoja, columnas)

        for fila in datos:
            hoja.append(fila)

        wb.save(ruta_guardado)
        os.startfile(ruta_guardado)
    except Exception as e:
        messagebox.showerror("Error de Exportación", str(e))